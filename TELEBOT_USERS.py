from GLOBAL_VAR import global_var
import openpyxl
import configparser
config = configparser.ConfigParser()
config.read(global_var())
file_path = str(config["TELEBOT_USERS"]["file_path"])

def user_id_list():
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    users = []
    for x in range(1,len(ws['A'])+1):
        if ws.cell(row=x,column=ws.max_column).value == 'Enable':
            users.append(ws.cell(row=x, column=1).value)
    wb.close()
    return users

def admin_user_id_list():
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    users = []
    for x in range(1,len(ws['A'])+1):
        if ws.cell(row=x,column=ws.max_column-1).value == 'y':
            users.append(ws.cell(row=x, column=1).value)
    wb.close()
    return users

def register_user(user_data):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    user_data.append('n')
    user_data.append('Request')
    ws.append(user_data)
    wb.save(file_path)
    wb.close()